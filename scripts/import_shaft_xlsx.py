from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / 'app' / 'modules' / 'shaft' / 'static' / 'data'
ELEMENTS = ['光', '灵', '咒', '暗', '魂', '相']

CHARACTER_IMAGE_OVERRIDES = {
    '主角': {
        'avatar': '/static/images/characters/avatar/鉴定师.webp',
        'portrait': '/static/images/characters/portrait/鉴定师.webp',
    },
    '真红': {
        'avatar': '/static/kongmu/images/characters/player_zhenhong_256.webp',
        'portrait': '/static/kongmu/images/characters/player_zhenhong_256.webp',
    },
    '伊洛伊': {
        'avatar': '/static/kongmu/images/characters/player_yiluoyi_256.webp',
        'portrait': '/static/kongmu/images/characters/player_yiluoyi_256.webp',
    },
}

SUBSTAT_UNITS = {
    'all_dmg': {'label': '通伤', 'unit_value': 0.01, 'kind': 'percent'},
    'crit_rate': {'label': '暴击', 'unit_value': 0.01, 'kind': 'percent'},
    'crit_dmg': {'label': '暴伤', 'unit_value': 0.02, 'kind': 'percent'},
    'harmony_strength': {'label': '环合强度', 'unit_value': 6, 'kind': 'flat'},
    'stagger_strength': {'label': '倾陷强度', 'unit_value': 6, 'kind': 'flat'},
    'atk_pct': {'label': '攻击%', 'unit_value': 0.0125, 'kind': 'percent'},
    'flat_atk': {'label': '攻击', 'unit_value': 8, 'kind': 'flat'},
    'hp_pct': {'label': '生命%', 'unit_value': 0.0125, 'kind': 'percent'},
    'flat_hp': {'label': '生命', 'unit_value': 100, 'kind': 'flat'},
    'def_pct': {'label': '防御%', 'unit_value': 0.0175, 'kind': 'percent'},
    'flat_def': {'label': '防御', 'unit_value': 8, 'kind': 'flat'},
}

CARTRIDGE_MAIN_STAT_OPTIONS = {
    '属伤': {'label': '属伤', 'modifier_key': 'element_dmg', 'unit_value': 0.375, 'kind': 'percent'},
    '暴击': {'label': '暴击', 'modifier_key': 'crit_rate', 'unit_value': 0.3, 'kind': 'percent'},
    '暴伤': {'label': '暴伤', 'modifier_key': 'crit_dmg', 'unit_value': 0.6, 'kind': 'percent'},
    '精通': {'label': '环合强度', 'modifier_key': 'harmony_strength', 'unit_value': 180, 'kind': 'flat'},
    '倾陷': {'label': '倾陷强度', 'modifier_key': 'stagger_strength', 'unit_value': 180, 'kind': 'flat'},
    '攻击': {'label': '攻击%', 'modifier_key': 'atk_pct', 'unit_value': 0.375, 'kind': 'percent'},
    '生命': {'label': '生命%', 'modifier_key': 'hp_pct', 'unit_value': 0.375, 'kind': 'percent'},
    '防御': {'label': '防御%', 'modifier_key': 'def_pct', 'unit_value': 0.525, 'kind': 'percent'},
    '治疗': {'label': '治疗', 'modifier_key': '', 'unit_value': 0, 'kind': 'flat'},
}

CURTAIN_BONUS_STAT_OPTIONS = {
    '属伤': {'label': '属伤', 'modifier_key': 'element_dmg', 'kind': 'percent'},
    '暴击': {'label': '暴击', 'modifier_key': 'crit_rate', 'kind': 'percent'},
    '暴伤': {'label': '暴伤', 'modifier_key': 'crit_dmg', 'kind': 'percent'},
    '充能': {'label': '充能', 'modifier_key': 'energy_recharge', 'kind': 'percent'},
    '攻击': {'label': '攻击%', 'modifier_key': 'atk_pct', 'kind': 'percent'},
    '生命': {'label': '生命%', 'modifier_key': 'hp_pct', 'kind': 'percent'},
    '防御': {'label': '防御%', 'modifier_key': 'def_pct', 'kind': 'percent'},
    '精通': {'label': '环合强度', 'modifier_key': 'harmony_strength', 'kind': 'percent'},
    '倾陷': {'label': '倾陷强度', 'modifier_key': 'stagger_strength', 'kind': 'percent'},
}

SUBSTAT_START_COLUMN = 36


def stable_id(prefix: str, name: str) -> str:
    digest = hashlib.sha1(name.encode('utf-8')).hexdigest()[:10]
    return f'{prefix}_{digest}'


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def web_text(value: Any) -> str:
    return str(normalize_value(value) or '').replace('精通', '环合强度').replace('穿防', '减防')


def number(value: Any, default: float = 0.0) -> float:
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def integer(value: Any, default: int = 0) -> int:
    return int(round(number(value, default)))


def get_row_map(ws: Any, row: int, max_col: int | None = None) -> dict[str, Any]:
    max_col = max_col or ws.max_column
    return {
        ws.cell(2, col).value: normalize_value(ws.cell(row, col).value)
        for col in range(1, max_col + 1)
        if ws.cell(2, col).value
    }


def image_for_character(name: str) -> dict[str, str]:
    if name in CHARACTER_IMAGE_OVERRIDES:
        return CHARACTER_IMAGE_OVERRIDES[name]
    avatar = ROOT / 'app' / 'static' / 'images' / 'characters' / 'avatar' / f'{name}.webp'
    portrait = ROOT / 'app' / 'static' / 'images' / 'characters' / 'portrait' / f'{name}.webp'
    return {
        'avatar': f'/static/images/characters/avatar/{name}.webp' if avatar.exists() else '',
        'portrait': f'/static/images/characters/portrait/{name}.webp' if portrait.exists() else '',
    }


def action_type_from_label(label: str) -> str:
    text = str(label or '').upper()
    if text in {'E', 'Q'}:
        return text
    if '援护' in text:
        return '援护'
    if '普攻' in text:
        return '普攻'
    if '心灵' in text:
        return '心灵'
    if text == '无':
        return '无'
    return str(label or '其他')


def action_resource_profile(action_type: str, time_seconds: float) -> dict[str, Any]:
    if action_type == 'E':
        return {'cooldown_ticks': max(1, int(round(max(time_seconds, 12.0) * 10))), 'energy_cost': 0}
    if action_type == 'Q':
        return {'cooldown_ticks': 0, 'energy_cost': 100}
    return {'cooldown_ticks': 0, 'energy_cost': 0}


def is_background_damage_action(action_name: Any, extra_tag: Any) -> bool:
    text = f'{action_name or ""} {extra_tag or ""}'
    return '后台' in text


def parse_bond_bonus(value: Any) -> dict[str, Any]:
    text = str(normalize_value(value) or '').strip()
    modifiers = {
        'crit_rate': 0.0,
        'atk_pct': 0.0,
        'hp_pct': 0.0,
        'def_pct': 0.0,
    }
    amount_text = ''.join(ch for ch in text if ch.isdigit() or ch == '.')
    amount = number(amount_text) / 100 if amount_text else 0.0
    if '暴击' in text:
        modifiers['crit_rate'] = amount
    elif '生命' in text:
        modifiers['hp_pct'] = amount
    elif '防御' in text:
        modifiers['def_pct'] = amount
    elif '攻击' in text:
        modifiers['atk_pct'] = amount
    return {
        'label': web_text(text),
        'modifiers': modifiers,
    }


def extract_characters(wb_values: Any) -> list[dict[str, Any]]:
    ws = wb_values['角色']
    build_ws = wb_values['配装']
    adaptation_by_name = {
        str(normalize_value(build_ws.cell(row, 1).value) or ''): normalize_value(build_ws.cell(row, 3).value) or ''
        for row in range(3, build_ws.max_row + 1)
        if normalize_value(build_ws.cell(row, 1).value)
    }
    bond_by_name = {
        str(normalize_value(build_ws.cell(row, 1).value) or ''): parse_bond_bonus(build_ws.cell(row, 57).value)
        for row in range(3, build_ws.max_row + 1)
        if normalize_value(build_ws.cell(row, 1).value)
    }
    records = []
    for row in range(3, ws.max_row + 1):
        name = normalize_value(ws.cell(row, 1).value)
        if not name:
            continue
        images = image_for_character(str(name))
        records.append({
            'id': stable_id('char', str(name)),
            'name': name,
            'element': normalize_value(ws.cell(row, 2).value) or '',
            'adaptation': adaptation_by_name.get(str(name), ''),
            'rarity': normalize_value(ws.cell(row, 3).value) or '',
            'level': 80,
            'avatar': images['avatar'],
            'portrait': images['portrait'],
            'base_stats': {
                'atk': number(ws.cell(row, 4).value),
                'hp': number(ws.cell(row, 6).value),
                'def': number(ws.cell(row, 8).value),
            },
            'modifiers': {
                # 角色表只提供基础攻击/生命/防御。角色被动与觉醒必须进入
                # buffs.json，不能把 xlsx 中预先揉入的面板增益重复导入。
                'crit_rate': 0.05,
                'crit_dmg': 0.5,
                'atk_pct': 0.0,
                'flat_atk': 0.0,
                'hp_pct': 0.0,
                'flat_hp': 0.0,
                'def_pct': 0.0,
                'flat_def': 0.0,
                'def_ignore': 0.0,
                'res_down': 0.0,
                'energy_recharge': 0.0,
                'harmony_strength': 0.0,
                'stagger_strength': 0.0,
                'basic_dmg': 0.0,
                'element_dmg': 0.0,
                'follow_dmg': 0.0,
                'mind_dmg': 0.0,
                'attach_dmg': 0.0,
                'skill_dmg': 0.0,
                'ultimate_dmg': 0.0,
                'all_dmg': 0.0,
            },
            'bond_bonus': bond_by_name.get(str(name), parse_bond_bonus('')),
            'source_row': row,
        })
    return records


def normalize_awakening_character_name(value: Any) -> str:
    name = str(normalize_value(value) or '').replace('「', '').replace('」', '').strip()
    if name in {'零', '鉴定师'}:
        return '主角'
    return name


def extract_awakenings(wb_values: Any) -> dict[str, list[dict[str, str]]]:
    if '角色觉醒' not in wb_values.sheetnames:
        return {}
    ws = wb_values['角色觉醒']
    awakenings: dict[str, list[dict[str, str]]] = {}
    seen: set[tuple[str, str, str]] = set()
    for row in range(2, ws.max_row + 1):
        character_name = normalize_awakening_character_name(ws.cell(row, 1).value)
        title = str(normalize_value(ws.cell(row, 2).value) or '')
        description = web_text(ws.cell(row, 3).value)
        if not character_name or not title or not description:
            continue
        identity = (character_name, title, description)
        if identity in seen:
            continue
        seen.add(identity)
        awakenings.setdefault(character_name, []).append({
            'title': title,
            'description': description,
        })
    return awakenings


def extract_arcs(wb_values: Any) -> list[dict[str, Any]]:
    ws = wb_values['弧盘']
    records = []
    for row in range(3, ws.max_row + 1):
        name = normalize_value(ws.cell(row, 1).value)
        if not name:
            continue
        element_bonus = {
            element: number(ws.cell(row, 27 + index).value)
            for index, element in enumerate(ELEMENTS)
        }
        records.append({
            'id': stable_id('arc', str(name)),
            'name': name,
            'level': 80,
            'rarity': normalize_value(ws.cell(row, 4).value) or '',
            'adaptation': normalize_value(ws.cell(row, 6).value) or '',
            'detail': web_text(ws.cell(row, 15).value),
            'base_atk': number(ws.cell(row, 2).value),
            'modifiers': {
                'crit_rate': number(ws.cell(row, 7).value) + number(ws.cell(row, 16).value),
                'crit_dmg': number(ws.cell(row, 8).value) + number(ws.cell(row, 17).value),
                'atk_pct': number(ws.cell(row, 9).value) + number(ws.cell(row, 20).value),
                'hp_pct': number(ws.cell(row, 10).value) + number(ws.cell(row, 21).value),
                'def_pct': number(ws.cell(row, 11).value) + number(ws.cell(row, 22).value),
                'energy_recharge': number(ws.cell(row, 12).value) + number(ws.cell(row, 23).value),
                'harmony_strength': number(ws.cell(row, 13).value) + number(ws.cell(row, 24).value),
                'stagger_strength': number(ws.cell(row, 14).value) + number(ws.cell(row, 25).value),
                'def_ignore': number(ws.cell(row, 18).value),
                'res_down': number(ws.cell(row, 19).value),
                'all_dmg': number(ws.cell(row, 26).value),
                'basic_dmg': number(ws.cell(row, 33).value),
                'dodge_counter_dmg': number(ws.cell(row, 34).value),
                'skill_dmg': number(ws.cell(row, 35).value),
                'ultimate_dmg': number(ws.cell(row, 36).value),
                'mind_dmg': number(ws.cell(row, 37).value),
                'follow_dmg': number(ws.cell(row, 38).value),
                'attach_dmg': number(ws.cell(row, 39).value),
            },
            'element_dmg': element_bonus,
            'source_row': row,
        })
    return records


def extract_cartridges(wb_values: Any) -> list[dict[str, Any]]:
    ws = wb_values['卡带']
    records = []
    for row in range(3, ws.max_row + 1):
        name = normalize_value(ws.cell(row, 1).value)
        if not name:
            continue
        records.append({
            'id': stable_id('cartridge', str(name)),
            'name': name,
            'detail': web_text(ws.cell(row, 10).value),
            'modifiers': {
                'element_dmg': number(ws.cell(row, 2).value),
                'atk_pct': number(ws.cell(row, 3).value) + number(ws.cell(row, 15).value),
                'hp_pct': number(ws.cell(row, 4).value) + number(ws.cell(row, 16).value),
                'def_pct': number(ws.cell(row, 5).value),
                'energy_recharge': number(ws.cell(row, 6).value),
                'harmony_strength': number(ws.cell(row, 7).value),
                'stagger_strength': number(ws.cell(row, 8).value),
                'mind_dmg': number(ws.cell(row, 9).value) + number(ws.cell(row, 22).value),
                'def_ignore': number(ws.cell(row, 11).value),
                'res_down': number(ws.cell(row, 12).value),
                'crit_rate': number(ws.cell(row, 13).value),
                'crit_dmg': number(ws.cell(row, 14).value),
                'all_dmg': number(ws.cell(row, 17).value),
                'basic_dmg': number(ws.cell(row, 19).value),
                'skill_dmg': number(ws.cell(row, 20).value),
                'ultimate_dmg': number(ws.cell(row, 21).value),
            },
            'shape_counts': {
                'type2': integer(ws.cell(row, 23).value),
                'type3': integer(ws.cell(row, 24).value),
                'type4': integer(ws.cell(row, 25).value),
            },
            'passive_counts': {
                'type2': integer(ws.cell(row, 26).value),
                'type3': integer(ws.cell(row, 27).value),
                'type4': integer(ws.cell(row, 28).value),
            },
            'source_row': row,
        })
    return records


def extract_actions(wb_values: Any, characters: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ws = wb_values['动作']
    char_by_name = {record['name']: record for record in characters}
    records = []
    for row in range(3, ws.max_row + 1):
        character_name = normalize_value(ws.cell(row, 1).value)
        action_name = normalize_value(ws.cell(row, 2).value)
        if not character_name or not action_name or character_name == '无':
            continue
        character = char_by_name.get(str(character_name))
        if not character:
            continue
        damage_type = normalize_value(ws.cell(row, 3).value) or '无'
        extra_tag = normalize_value(ws.cell(row, 4).value) or ''
        time_seconds = number(ws.cell(row, 9).value)
        duration_ticks = max(0, int(round(time_seconds * 10)))
        action_type = action_type_from_label(str(damage_type))
        resources = action_resource_profile(action_type, time_seconds)
        background_damage = is_background_damage_action(action_name, extra_tag)
        passive_multipliers = {
            ('主角', 'e额外'): 2.0,
            ('九原', '创生追加'): 0.15,
            ('安魂曲', '失谐强化'): 4.0,
            ('达芙蒂尔', '移行附加'): 3.0,
            ('海月', '黯星追加'): 1.5,
            ('卡厄斯', '延滞结算'): 8.0,
            ('卡厄斯', '哈索尔延滞'): 32.0,
            ('伊洛伊', 'B觉'): 1.5,
        }
        passive_multiplier = passive_multipliers.get((str(character_name), str(action_name)))
        if passive_multiplier is not None:
            action_type = '被动'
            damage_type = '被动'
            background_damage = True
            time_seconds = 0.0
            duration_ticks = 0
            resources = {'cooldown_ticks': 50 if (character_name, action_name) == ('伊洛伊', 'B觉') else 0, 'energy_cost': 0}
        elif str(character_name) == '伊洛伊' and str(action_name) in {'援护', 'e'}:
            time_seconds = 1.5
            duration_ticks = 15
        if str(character_name) == '伊洛伊' and str(action_name) == 'q持续':
            background_damage = True
            resources['energy_cost'] = 0
        if (str(character_name), str(action_name)) == ('真红', '龙e'):
            resources['energy_cost'] = 12
        imported_extra_tag = '' if extra_tag == '0' else str(extra_tag)
        if passive_multiplier is not None:
            imported_extra_tag = '协攻' if (character_name, action_name) == ('伊洛伊', 'B觉') else '附着'
        self_modifiers = {
            'crit_rate': number(ws.cell(row, 15).value),
            'crit_dmg': number(ws.cell(row, 16).value),
            'all_dmg': number(ws.cell(row, 17).value),
            'flat_atk': number(ws.cell(row, 18).value),
            'atk_pct': number(ws.cell(row, 19).value),
            'hp_pct': number(ws.cell(row, 20).value),
            'def_pct': number(ws.cell(row, 21).value),
            'harmony_strength': number(ws.cell(row, 22).value),
            'def_ignore': number(ws.cell(row, 23).value),
            'res_down': number(ws.cell(row, 24).value),
        }
        # These values are now expressed as auditable character buff rules.
        if (character_name, action_name) == ('主角', 'q'):
            self_modifiers['crit_rate'] = 0.0
        if (character_name, action_name) == ('达芙蒂尔', '移行'):
            self_modifiers['all_dmg'] = 0.0
        if (character_name, action_name) == ('海月', 'q'):
            self_modifiers['crit_rate'] = 0.0
            self_modifiers['all_dmg'] = 0.0
        records.append({
            'id': stable_id('action', f'{character_name}:{action_name}:{row}'),
            'character_id': character['id'],
            'character_name': character_name,
            'name': action_name,
            'action_type': action_type,
            'damage_type': damage_type,
            'extra_tag': imported_extra_tag,
            'is_background_damage': background_damage,
            'duration_seconds': time_seconds,
            'duration_ticks': duration_ticks,
            'multipliers': {
                'atk': passive_multiplier if passive_multiplier is not None else number(ws.cell(row, 5).value) / 100,
                'hp': number(ws.cell(row, 6).value) / 100,
                'def': number(ws.cell(row, 7).value) / 100,
                'flat': number(ws.cell(row, 8).value),
            },
            'energy_gain': number(ws.cell(row, 10).value),
            'harmony': number(ws.cell(row, 11).value),
            'stagger': number(ws.cell(row, 12).value),
            'energy_return': number(ws.cell(row, 14).value),
            'self_modifiers': self_modifiers,
            'cooldown_ticks': resources['cooldown_ticks'],
            'energy_cost': resources['energy_cost'],
            'personal_resource_cost': {},
            'personal_resource_gain': {},
            'required_awakening': 2 if (character_name, action_name) == ('伊洛伊', 'B觉') else 0,
            'source_row': row,
        })
    return records


def name_to_id_map(records: list[dict[str, Any]]) -> dict[str, str]:
    return {str(record['name']): str(record['id']) for record in records}


def extract_substat_counts(build_ws: Any) -> dict[str, dict[str, int]]:
    counts_by_character: dict[str, dict[str, int]] = {}
    keys = list(SUBSTAT_UNITS)
    for row in range(3, build_ws.max_row + 1):
        character_name = normalize_value(build_ws.cell(row, 1).value)
        if not character_name:
            continue
        counts: dict[str, int] = {}
        for index, key in enumerate(keys):
            unit = number(SUBSTAT_UNITS[key].get('unit_value'), 1)
            value = number(build_ws.cell(row, SUBSTAT_START_COLUMN + index).value)
            counts[key] = max(0, int(round(value / unit))) if unit else 0
        counts_by_character[str(character_name)] = counts
    return counts_by_character


def passive_type_from_label(value: Any) -> str:
    text = str(normalize_value(value) or '')
    if '2' in text:
        return 'type2'
    if '4' in text:
        return 'type4'
    return 'type3'


def extract_default_build_options(build_ws: Any, character_ids: dict[str, str]) -> dict[str, dict[str, Any]]:
    defaults: dict[str, dict[str, Any]] = {}
    for row in range(3, build_ws.max_row + 1):
        character_name = normalize_value(build_ws.cell(row, 1).value)
        character_id = character_ids.get(str(character_name), '')
        if not character_id:
            continue
        defaults[character_id] = {
            'cartridge_main_stat': normalize_value(build_ws.cell(row, 30).value) or '',
            'curtain_bonus': {
                'value': number(build_ws.cell(row, 31).value),
                'stat': normalize_value(build_ws.cell(row, 32).value) or '',
                'passive_type': passive_type_from_label(build_ws.cell(row, 34).value),
            },
        }
    return defaults


def extract_character_builds(
    build_ws: Any,
    character_ids: dict[str, str],
    arc_ids: dict[str, str],
    cartridge_ids: dict[str, str],
    substat_counts_by_character: dict[str, dict[str, int]],
) -> dict[str, dict[str, Any]]:
    builds: dict[str, dict[str, Any]] = {}
    for row in range(3, build_ws.max_row + 1):
        character_name = normalize_value(build_ws.cell(row, 1).value)
        if not character_name:
            continue
        character_id = character_ids.get(str(character_name), '')
        if not character_id:
            continue
        arc_name = normalize_value(build_ws.cell(row, 24).value) or ''
        cartridge_name = normalize_value(build_ws.cell(row, 26).value) or ''
        bond_full = bool(build_ws.cell(row, 56).value)
        builds[character_id] = {
            'character_id': character_id,
            'character_name': character_name,
            'arc_id': arc_ids.get(str(arc_name), ''),
            'arc_name': arc_name,
            'cartridge_id': cartridge_ids.get(str(cartridge_name), ''),
            'cartridge_name': cartridge_name,
            'awakening': integer(build_ws.cell(row, 53).value),
            'bond_level': 1 if bond_full else 0,
            'bond_full': bond_full,
            'cartridge_main_stat': normalize_value(build_ws.cell(row, 30).value) or '',
            'curtain_bonus': {
                'value': number(build_ws.cell(row, 31).value),
                'stat': normalize_value(build_ws.cell(row, 32).value) or '',
                'passive_type': passive_type_from_label(build_ws.cell(row, 34).value),
            },
            'substat_counts': substat_counts_by_character.get(str(character_name), {key: 0 for key in SUBSTAT_UNITS}),
        }
    return builds


def extract_starter_axis(wb_values: Any, characters: list[dict[str, Any]], arcs: list[dict[str, Any]], cartridges: list[dict[str, Any]], actions: list[dict[str, Any]]) -> dict[str, Any]:
    ws = wb_values['云配队']
    build_ws = wb_values['配装']
    character_ids = name_to_id_map(characters)
    arc_ids = name_to_id_map(arcs)
    cartridge_ids = name_to_id_map(cartridges)
    substat_counts_by_character = extract_substat_counts(build_ws)
    actions_by_pair = {
        (str(action['character_name']), str(action['name'])): action
        for action in actions
    }
    character_builds = extract_character_builds(
        build_ws,
        character_ids,
        arc_ids,
        cartridge_ids,
        substat_counts_by_character,
    )

    team = []
    for slot, row in enumerate(range(4, 8)):
        character_name = normalize_value(ws.cell(row, 6).value) or ''
        arc_name = normalize_value(ws.cell(row, 7).value) or ''
        cartridge_name = normalize_value(ws.cell(row, 8).value) or ''
        if not character_name:
            continue
        team.append({
            'slot': slot,
            'character_id': character_ids.get(str(character_name), ''),
            'character_name': character_name,
            'arc_id': arc_ids.get(str(arc_name), ''),
            'arc_name': arc_name,
            'cartridge_id': cartridge_ids.get(str(cartridge_name), ''),
            'cartridge_name': cartridge_name,
            'awakening': integer(ws.cell(row, 9).value),
            'substat_counts': substat_counts_by_character.get(str(character_name), {key: 0 for key in SUBSTAT_UNITS}),
        })

    steps = []
    current_tick = 0
    for col in range(10, 30):
        for slot, row in enumerate(range(4, 8)):
            action_name = normalize_value(ws.cell(row, col).value)
            if not action_name or action_name == '0':
                continue
            character_name = team[slot]['character_name'] if slot < len(team) else ''
            action = actions_by_pair.get((str(character_name), str(action_name)))
            if action is None:
                continue
            steps.append({
                'id': f'step_{len(steps) + 1:03d}',
                'slot': slot,
                'action_id': action['id'],
                'action_name': action['name'],
                'start_tick': current_tick,
                'repeat': 1,
                'tags': [],
            })
            current_tick += max(1, int(action.get('duration_ticks') or 0))

    return {
        'axis_version': 1,
        'title': 'xlsx 示例轴',
        'duration_ticks': max(current_tick, 600),
        'team': team,
        'character_builds': character_builds,
        'steps': steps[:30],
        'enemy': {
            'level': integer(ws.cell(9, 7).value, 90),
            'track_outside': bool(ws.cell(13, 6).value),
            'weakness_elements': [str(ws.cell(11, 7).value)] if ws.cell(11, 7).value else [],
        },
        'options': {
            'reaction_mastery_policy': 'team_max',
            'crit_mode': 'expected',
            'switch_loss_ticks': 2,
            'front_state_enabled': True,
        },
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8') as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write('\n')


def main() -> int:
    parser = argparse.ArgumentParser(description='Import shaft calculator data from 异环云配队.xlsx.')
    parser.add_argument('xlsx_path', type=Path)
    parser.add_argument('--output-dir', type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    source_bytes = args.xlsx_path.read_bytes()
    source_hash = hashlib.sha256(source_bytes).hexdigest()
    wb_values = openpyxl.load_workbook(args.xlsx_path, data_only=True, read_only=False)

    characters = extract_characters(wb_values)
    awakenings = extract_awakenings(wb_values)
    arcs = extract_arcs(wb_values)
    cartridges = extract_cartridges(wb_values)
    actions = extract_actions(wb_values, characters)
    energy_capacity_by_character: dict[str, float] = {}
    for action in actions:
        if action.get('action_type') != 'Q' and action.get('damage_type') != 'Q':
            continue
        character_id = str(action.get('character_id') or '')
        energy_capacity_by_character[character_id] = max(
            energy_capacity_by_character.get(character_id, 0.0),
            number(action.get('energy_cost')),
        )
    for character in characters:
        character['energy_capacity'] = energy_capacity_by_character.get(str(character.get('id') or ''), 0.0)
    starter_axis = extract_starter_axis(wb_values, characters, arcs, cartridges, actions)

    source_meta = {
        'source_name': args.xlsx_path.name,
        'source_hash': source_hash,
        'imported_at': datetime.now(timezone.utc).isoformat(),
        'version_label': '异环云配队 1.0.0',
        'sheets': {
            sheet_name: {
                'rows': wb_values[sheet_name].max_row,
                'cols': wb_values[sheet_name].max_column,
            }
            for sheet_name in wb_values.sheetnames
        },
        'notes': [
            '运行时只读取项目内静态 JSON，不读取 xlsx。',
            'xlsx 原“精通”字段在网页中显示为“环合强度”。',
        ],
    }

    formula_constants = {
        'substat_units': SUBSTAT_UNITS,
        'cartridge_main_stat_options': CARTRIDGE_MAIN_STAT_OPTIONS,
        'curtain_bonus_stat_options': CURTAIN_BONUS_STAT_OPTIONS,
        'default_build_options': extract_default_build_options(wb_values['配装'], name_to_id_map(characters)),
        'elements': ELEMENTS,
        'default_enemy': {
            'level': 90,
            'track_outside': False,
            'weakness_elements': ['光'],
        },
        'switch_loss_ticks': 2,
        'damage_formula': 'base * (1 + dmg_bonus) * crit_expected * resistance * defense',
    }

    write_json(args.output_dir / 'source_meta.json', source_meta)
    write_json(args.output_dir / 'characters.json', characters)
    if awakenings:
        write_json(args.output_dir / 'awakenings.json', awakenings)
    write_json(args.output_dir / 'arcs.json', arcs)
    write_json(args.output_dir / 'cartridges.json', cartridges)
    write_json(args.output_dir / 'actions.json', actions)
    write_json(args.output_dir / 'starter_axis.json', starter_axis)
    write_json(args.output_dir / 'formula_constants.json', formula_constants)
    print(f'imported shaft data to {args.output_dir}')
    print(f'characters={len(characters)} arcs={len(arcs)} cartridges={len(cartridges)} actions={len(actions)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
